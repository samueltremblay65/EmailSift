from openpyxl import load_workbook
from simplegmail.query import construct_query

def open_workbook(path):
    workbook = load_workbook(filename=path)
    return workbook

def delete_from_excel_sheet(excel_file, gmail):

    # Obtain data from excel sheet
    workbook = open_workbook(excel_file)

    all_deletions = []

    for worksheet in workbook.sheetnames:
        sheet = workbook[worksheet]
        deletions = process_sheet(gmail, sheet)

        for deletion in deletions:
            if not any(email.id == deletion.id for email in all_deletions):
                all_deletions.append(deletion)

    for deletion in all_deletions:
        deletion.trash()

    return len(all_deletions)

def process_sheet(gmail, sheet):
    deletion_candidates = []

    # Get from list
    from_list = []
    i = 2
    while sheet.cell(i, 1).value is not None:
        from_list.append(sheet.cell(i, 1).value)
        i += 1

    # Get delete list
    keep_list = []
    i = 2
    while sheet.cell(i, 2).value is not None:
        keep_list.append(sheet.cell(i, 2).value.lower())
        i += 1

    # Get keep list
    delete_list = []
    i = 2
    while sheet.cell(i, 3).value is not None:
        delete_list.append(sheet.cell(i, 3).value.lower())
        i += 1

    # Get keep senders
    keep_senders = []
    i = 2
    while sheet.cell(i, 4).value is not None:
        keep_senders.append(sheet.cell(i, 4).value.lower())
        i += 1

    query_params = {
        "sender": from_list
    }

    # Get list of emails affected by the query
    if("*" in from_list):
        emails = gmail.get_messages()
    else:
        emails = gmail.get_messages(query=construct_query(query_params))

    # Treating special cases
    if("*" in delete_list):
        if("first" in keep_list or "First" in keep_list):
            return emails[1:]
        if("last" in keep_list or "Last" in keep_list):
            return emails[:1]
        return emails

    deletion_candidates = []

    for email in emails:
        for word in delete_list:
            if not (email.html is None) and word in email.html.lower():
                deletion_candidates.append(email)
                break
            if not (email.plain is None) and word in email.plain.lower():
                deletion_candidates.append(email)
                break
            if not (email.subject is None) and word in email.subject.lower():
                deletion_candidates.append(email)
                break

    # Remove items that are specified to be kept from the deletion candidates
    deletions = apply_keep_logic(deletion_candidates, keep_list, keep_senders)

    return deletions

def apply_keep_logic(candidates, keep_list, keep_senders):
    deletions = [email for email in candidates if not contains_keep_word(email, keep_list) and not from_keep_sender(email, keep_senders)]
    return deletions

def contains_keep_word(email, keep_list):
    for word in keep_list:
        if not (email.html is None) and word in email.html:
            return True
        elif not (email.plain is None) and word in email.plain:
            return True
        elif not (email.subject is None) and word in email.subject:
            return True
    return False

def from_keep_sender(email, keep_senders):
    if(email.sender in keep_senders):
        return True
    return False
